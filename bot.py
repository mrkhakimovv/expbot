import asyncio
import os
import json
import logging
from aiogram import Bot, Dispatcher, types, F
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.filters import Command
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from openpyxl import Workbook, load_workbook

# Log sozlash
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Bot token
API_TOKEN = "8247447507:AAF2V-7ET2l7D3eDFJ3Wv4IvJrdaHxTWUJg"
ADMIN_ID = 1986422890  # Asosiy admin ID

# Bot va Dispatcher yaratish
bot = Bot(token=API_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher()

# Fayllar
EXCEL_FILE = "users.xlsx"
ADMINS_FILE = "admins.json"
CHANNELS_FILE = "channels.json"

# Ma'lumotlarni yuklash funksiyalari
def load_admins():
    if os.path.exists(ADMINS_FILE):
        with open(ADMINS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return [ADMIN_ID]  # Asosiy admin

def save_admins(admins):
    with open(ADMINS_FILE, 'w', encoding='utf-8') as f:
        json.dump(admins, f, ensure_ascii=False, indent=2)

def load_channels():
    if os.path.exists(CHANNELS_FILE):
        with open(CHANNELS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []  # Boshlang'ich kanallar ro'yxati

def save_channels(channels):
    with open(CHANNELS_FILE, 'w', encoding='utf-8') as f:
        json.dump(channels, f, ensure_ascii=False, indent=2)

# Excel faylni yaratish
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Foydalanuvchilar"
        ws.append(["User ID", "Ism Familiya", "Telefon raqam", "Ro'yxatdan o'tgan sana"])
        wb.save(EXCEL_FILE)
        logger.info("Excel fayli yaratildi")

# Ma'lumotlarni yuklash
admins_list = load_admins()
channels_list = load_channels()

# Vaqtincha ma'lumotlarni saqlash
user_temp = {}
admin_temp = {}

# Admin tekshirish
def is_admin(user_id: int) -> bool:
    return user_id in admins_list

# Kanalga obuna tekshirish
async def check_subscription(user_id: int) -> bool:
    if not channels_list:
        return True  # Agar kanal qo'shilmagan bo'lsa, tekshirmaymiz
    
    for channel in channels_list:
        try:
            member = await bot.get_chat_member(chat_id=channel['id'], user_id=user_id)
            if member.status not in ["member", "administrator", "creator"]:
                return False
        except Exception as e:
            logger.error(f"Kanal tekshirishda xato: {e}")
            continue
    return True

# Admin paneli
def get_admin_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📊 Statistika"), KeyboardButton(text="👥 Foydalanuvchilar")],
            [KeyboardButton(text="📢 Reklama jo'natish"), KeyboardButton(text="➕ Admin qo'shish")],
            [KeyboardButton(text="📺 Kanallar boshqaruvi"), KeyboardButton(text="🔙 Asosiy menyu")]
        ],
        resize_keyboard=True,
        input_field_placeholder="Admin paneli..."
    )

# Kanallar boshqaruvi keyboard
def get_channels_keyboard():
    keyboard = []
    for channel in channels_list:
        keyboard.append([InlineKeyboardButton(
            text=f"❌ {channel['title']}", 
            callback_data=f"delete_channel:{channel['id']}"
        )])
    
    keyboard.append([InlineKeyboardButton(text="➕ Kanal qo'shish", callback_data="add_channel")])
    keyboard.append([InlineKeyboardButton(text="🔙 Orqaga", callback_data="back_to_admin")])
    
    return InlineKeyboardMarkup(inline_keyboard=keyboard)

# START komandasi
@dp.message(Command("start"))
async def start_cmd(message: types.Message):
    user_id = message.from_user.id
    user_name = message.from_user.full_name
    
    logger.info(f"Start bosildi: {user_id} - {user_name}")
    
    # Admin tekshirish
    if is_admin(user_id):
        await message.answer(
            f"👋 Admin paneliga xush kelibsiz, {user_name}!",
            reply_markup=get_admin_keyboard()
        )
        return
    
    # Obunani tekshirish
    if not await check_subscription(user_id):
        # Kanallar ro'yxatini tayyorlash
        channels_text = ""
        keyboard_buttons = []
        
        for channel in channels_list:
            channels_text += f"• {channel['title']}\n"
            keyboard_buttons.append([
                InlineKeyboardButton(
                    text=f"📢 {channel['title']}", 
                    url=channel['invite_link']
                )
            ])
        
        keyboard_buttons.append([InlineKeyboardButton(text="✅ Obunani tekshirish", callback_data="check_subs")])
        
        await message.answer(
            f"🤖 Botdan to'liq foydalanish uchun quyidagi kanallarga obuna bo'ling:\n\n{channels_text}",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=keyboard_buttons)
        )
        return

    # Agar obuna bo'lsa
    await message.answer(
        f"Assalomu alaykum {user_name}! 🎉\n\n"
        "Botimizga xush kelibsiz! Ro'yxatdan o'tish uchun quyidagi amallarni bajaring:\n\n"
        "1. Ism va familiyangizni yuboring:"
    )
    user_temp[user_id] = {"step": "name"}

# Obunani tekshirish
@dp.callback_query(F.data == "check_subs")
async def check_subs(callback: types.CallbackQuery):
    user_id = callback.from_user.id
    
    if await check_subscription(user_id):
        await callback.message.edit_text(
            "✅ Barcha kanallarga obuna tasdiqlandi! Endi ro'yxatdan o'tishni davom ettiramiz.\n\n"
            "Ism va familiyangizni yuboring:"
        )
        user_temp[user_id] = {"step": "name"}
    else:
        await callback.answer("❌ Hali barcha kanallarga obuna bo'lmagansiz!", show_alert=True)

# Ism familiya qabul qilish
@dp.message(F.text & ~F.via_bot)
async def get_name(message: types.Message):
    user_id = message.from_user.id
    
    # Admin paneli buyruqlari
    if is_admin(user_id):
        if message.text == "📊 Statistika":
            await show_stats(message)
            return
        elif message.text == "👥 Foydalanuvchilar":
            await send_excel(message)
            return
        elif message.text == "📢 Reklama jo'natish":
            await request_advertisement(message)
            return
        elif message.text == "➕ Admin qo'shish":
            await request_new_admin(message)
            return
        elif message.text == "📺 Kanallar boshqaruvi":
            await manage_channels(message)
            return
        elif message.text == "🔙 Asosiy menyu":
            await message.answer("Asosiy menyu", reply_markup=types.ReplyKeyboardRemove())
            return
    
    # Foydalanuvchi ro'yxatdan o'tish jarayoni
    if user_id in user_temp and user_temp[user_id].get("step") == "name":
        if len(message.text.strip()) < 3:
            await message.answer("❌ Iltimos, to'liq ism va familiyangizni kiriting (kamida 3 ta belgi):")
            return
            
        user_temp[user_id] = {
            "step": "phone",
            "name": message.text.strip()
        }

        kb = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="📱 Telefon raqamni ulashish", request_contact=True)]
            ],
            resize_keyboard=True,
            one_time_keyboard=True
        )
        
        await message.answer(
            f"✅ Rahmat! Endi telefon raqamingizni yuboring:",
            reply_markup=kb
        )
    
    elif user_id not in user_temp and not is_admin(user_id):
        await message.answer("Botni ishga tushirish uchun /start buyrug'ini bosing.")

# Telefon raqam qabul qilish
@dp.message(F.contact)
async def get_contact(message: types.Message):
    user_id = message.from_user.id
    
    if user_id in user_temp and user_temp[user_id].get("step") == "phone":
        name = user_temp[user_id]["name"]
        phone = message.contact.phone_number
        
        # Excelga ma'lumotlarni yozish
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            
            from datetime import datetime
            registration_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            ws.append([user_id, name, phone, registration_date])
            wb.save(EXCEL_FILE)
            
            logger.info(f"Yangi foydalanuvchi qo'shildi: {user_id} - {name}")
            
            await message.answer(
                f"🎉 Tabriklaymiz! Ro'yxatdan muvaffaqiyatli o'tdingiz!\n\n"
                f"Endi botning barcha imkoniyatlaridan foydalanishingiz mumkin.",
                reply_markup=types.ReplyKeyboardRemove()
            )
            
            # Adminga bildirishnoma
            for admin_id in admins_list:
                try:
                    await bot.send_message(
                        admin_id,
                        f"🆕 Yangi foydalanuvchi:\nID: {user_id}\nIsm: {name}\nTel: {phone}"
                    )
                except Exception as e:
                    logger.error(f"Adminga xabar yuborishda xato: {e}")
            
            user_temp.pop(user_id)
            
        except Exception as e:
            logger.error(f"Excelga yozishda xato: {e}")
            await message.answer(
                "❌ Ma'lumotlarni saqlashda xatolik. Iltimos, qaytadan urinib ko'ring.",
                reply_markup=types.ReplyKeyboardRemove()
            )
            user_temp.pop(user_id)
    
    else:
        await message.answer(
            "Iltimos, avval ism va familiyangizni yuboring.",
            reply_markup=types.ReplyKeyboardRemove()
        )

# ==================== ADMIN FUNCTIONS ====================

# Statistika
async def show_stats(message: types.Message):
    try:
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            user_count = ws.max_row - 1
            
            await message.answer(
                f"📈 Bot statistikasi:\n\n"
                f"• Jami foydalanuvchilar: {user_count} ta\n"
                f"• Adminlar soni: {len(admins_list)} ta\n"
                f"• Majburiy kanallar: {len(channels_list)} ta\n"
                f"• Faol jarayondagilar: {len(user_temp)} ta"
            )
        else:
            await message.answer("❌ Excel fayli topilmadi.")
    except Exception as e:
        logger.error(f"Statistikani ko'rsatishda xato: {e}")
        await message.answer("❌ Statistikani yuklashda xatolik.")

# Foydalanuvchilar ro'yxati
async def send_excel(message: types.Message):
    try:
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            user_count = ws.max_row - 1
            
            await message.answer_document(
                types.FSInputFile(EXCEL_FILE),
                caption=f"📊 Foydalanuvchilar ro'yxati\nJami: {user_count} ta foydalanuvchi"
            )
        else:
            await message.answer("❌ Excel fayli topilmadi.")
    except Exception as e:
        logger.error(f"Excel faylini yuborishda xato: {e}")
        await message.answer("❌ Faylni yuborishda xatolik yuz berdi.")

# Reklama so'rash
async def request_advertisement(message: types.Message):
    admin_temp[message.from_user.id] = {"action": "advertisement"}
    await message.answer(
        "📢 Reklama postini yuboring (matn, rasm, video yoki hujjat):\n\n"
        "⚠️ Eslatma: Bu xabar BARCHA foydalanuvchilarga yuboriladi!",
        reply_markup=ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="🔙 Bekor qilish")]],
            resize_keyboard=True
        )
    )

# Yangi admin so'rash
async def request_new_admin(message: types.Message):
    admin_temp[message.from_user.id] = {"action": "add_admin"}
    await message.answer(
        "Yangi adminning User ID sini yuboring:\n\n"
        "🆔 User ID ni olish uchun @userinfobot dan foydalanishingiz mumkin",
        reply_markup=ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="🔙 Bekor qilish")]],
            resize_keyboard=True
        )
    )

# Kanallar boshqaruvi
async def manage_channels(message: types.Message):
    if not channels_list:
        text = "📺 Hozircha hech qanday kanal qo'shilmagan"
    else:
        text = f"📺 Majburiy obuna kanallari ({len(channels_list)} ta):"
    
    await message.answer(
        text,
        reply_markup=get_channels_keyboard()
    )

# Kanal qo'shish
@dp.callback_query(F.data == "add_channel")
async def add_channel_callback(callback: types.CallbackQuery):
    admin_temp[callback.from_user.id] = {"action": "add_channel"}
    await callback.message.edit_text(
        "Yangi kanalni quyidagi formatda yuboring:\n\n"
        "Kanal nomi\n"
        "@kanal_username\n"
        "https://t.me/kanal_username\n\n"
        "Misol:\n"
        "Hakimov Math\n"
        "@hakimov_math\n"
        "https://t.me/hakimov_math"
    )

# Kanal o'chirish
@dp.callback_query(F.data.startswith("delete_channel:"))
async def delete_channel_callback(callback: types.CallbackQuery):
    channel_id = int(callback.data.split(":")[1])
    
    global channels_list
    channels_list = [ch for ch in channels_list if ch['id'] != channel_id]
    save_channels(channels_list)
    
    await callback.message.edit_text(
        "✅ Kanal o'chirildi!",
        reply_markup=get_channels_keyboard()
    )
    await callback.answer("Kanal o'chirildi")

# Admin paneliga qaytish
@dp.callback_query(F.data == "back_to_admin")
async def back_to_admin(callback: types.CallbackQuery):
    await callback.message.edit_text("Admin paneli:")
    await callback.message.answer(
        "Admin paneli",
        reply_markup=get_admin_keyboard()
    )

# Reklama jo'natish
async def send_advertisement_to_all(message: types.Message, content_type: str):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        user_ids = []
        for row in range(2, ws.max_row + 1):  # 1-qator sarlavha
            user_ids.append(ws.cell(row=row, column=1).value)
        
        success_count = 0
        fail_count = 0
        
        for user_id in user_ids:
            try:
                if content_type == "text":
                    await bot.send_message(user_id, message.text)
                elif content_type == "photo":
                    await bot.send_photo(user_id, message.photo[-1].file_id, caption=message.caption)
                elif content_type == "video":
                    await bot.send_video(user_id, message.video.file_id, caption=message.caption)
                elif content_type == "document":
                    await bot.send_document(user_id, message.document.file_id, caption=message.caption)
                
                success_count += 1
                await asyncio.sleep(0.1)  # Spamdan saqlanish uchun
                
            except Exception as e:
                fail_count += 1
                logger.error(f"Foydalanuvchiga xabar yuborishda xato: {user_id} - {e}")
        
        await message.answer(
            f"📊 Reklama jo'natish natijasi:\n\n"
            f"✅ Muvaffaqiyatli: {success_count} ta\n"
            f"❌ Xatolik: {fail_count} ta\n"
            f"📨 Jami: {len(user_ids)} ta",
            reply_markup=get_admin_keyboard()
        )
        
    except Exception as e:
        logger.error(f"Reklama jo'natishda xato: {e}")
        await message.answer("❌ Reklama jo'natishda xatolik yuz berdi.")

# Admin xabarlarini qayta ishlash
@dp.message(F.content_type.in_({types.ContentType.TEXT, types.ContentType.PHOTO, 
                               types.ContentType.VIDEO, types.ContentType.DOCUMENT}))
async def handle_admin_actions(message: types.Message):
    user_id = message.from_user.id
    
    if not is_admin(user_id) or user_id not in admin_temp:
        return
    
    action = admin_temp[user_id].get("action")
    
    if message.text == "🔙 Bekor qilish":
        admin_temp.pop(user_id)
        await message.answer("Amal bekor qilindi", reply_markup=get_admin_keyboard())
        return
    
    if action == "advertisement":
        content_type = "text"
        if message.photo:
            content_type = "photo"
        elif message.video:
            content_type = "video"
        elif message.document:
            content_type = "document"
        
        await send_advertisement_to_all(message, content_type)
        admin_temp.pop(user_id)
        
    elif action == "add_admin":
        try:
            new_admin_id = int(message.text)
            if new_admin_id not in admins_list:
                admins_list.append(new_admin_id)
                save_admins(admins_list)
                await message.answer(
                    f"✅ Yangi admin qo'shildi: {new_admin_id}",
                    reply_markup=get_admin_keyboard()
                )
            else:
                await message.answer("❌ Bu admin allaqachon mavjud")
        except ValueError:
            await message.answer("❌ Noto'g'ri ID format. Faqat raqam kiriting.")
        admin_temp.pop(user_id)
        
    elif action == "add_channel":
        try:
            lines = message.text.split('\n')
            if len(lines) >= 3:
                title = lines[0].strip()
                username = lines[1].strip().lstrip('@')
                invite_link = lines[2].strip()
                
                # Kanal ID sini olish
                try:
                    chat = await bot.get_chat(f"@{username}")
                    channel_id = chat.id
                    
                    new_channel = {
                        "id": channel_id,
                        "title": title,
                        "username": username,
                        "invite_link": invite_link
                    }
                    
                    channels_list.append(new_channel)
                    save_channels(channels_list)
                    
                    await message.answer(
                        f"✅ Kanal qo'shildi: {title}",
                        reply_markup=get_admin_keyboard()
                    )
                    
                except Exception as e:
                    await message.answer(f"❌ Kanal topilmadi yoki bot admin emas: {e}")
            else:
                await message.answer("❌ Noto'g'ri format. Qaytadan urinib ko'ring.")
        except Exception as e:
            await message.answer(f"❌ Xatolik: {e}")
        admin_temp.pop(user_id)

# Yordam komandasi
@dp.message(Command("help"))
async def help_cmd(message: types.Message):
    if is_admin(message.from_user.id):
        await message.answer(
            "🤖 Admin yordami:\n\n"
            "📊 Statistika - Bot statistikasi\n"
            "👥 Foydalanuvchilar - Excel faylini olish\n"
            "📢 Reklama - Barcha foydalanuvchilarga reklama jo'natish\n"
            "➕ Admin - Yangi admin qo'shish\n"
            "📺 Kanallar - Majburiy kanallarni boshqarish"
        )
    else:
        await message.answer(
            "🤖 Bot yordami:\n\n"
            "/start - Botni ishga tushirish\n"
            "/help - Yordam ko'rsatish"
        )

# Botni ishga tushirish
async def main():
    # Excel faylni ishga tushirish
    init_excel()
    
    logger.info("Bot ishga tushmoqda...")
    logger.info(f"Adminlar: {admins_list}")
    logger.info(f"Kanallar: {channels_list}")
    
    await bot.delete_webhook(drop_pending_updates=True)
    await dp.start_polling(bot)

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("Bot to'xtatildi")
    except Exception as e:
        logger.error(f"Botda xatolik: {e}")